import numpy as np
import streamlit as st
from openpyxl import load_workbook

st.set_page_config(layout="wide", page_title="Swara Pattern Calculator")

# Load the Excel file from GitHub repo
@st.cache_data
def load_data():
    workbook = load_workbook(filename="Corvai.xlsx", data_only=True)
    return workbook.active

sheet = load_data()

# Sidebar inputs
with st.sidebar:
    lang = st.selectbox('Select Language of Terminology', ["Telugu", "Tamil"])
    common = st.selectbox('Choose the Talam', ["Adi", "Rupakam", "Adi Chowkam", "Misra Chapu", "Khanda Chapu", "Adi (Tisra Gathi)", "Sapta Talas", "Other"])
    cycle = 0
    tala = common
    jathi = 0
    k = 0
    
    if common in [sheet.cell(row=i+5, column=1).value for i in range(6)]:
        cycle = next(sheet.cell(row=i+5, column=2).value for i in range(6) if sheet.cell(row=i+5, column=1).value == common)
    
    if common == "Sapta Talas":
        tala = st.selectbox('Choose the Talam', ["Dhruva", "Matya", "Rupaka", "Jhampa", "Triputa", "Ata", "Eka"])
        jathi = st.selectbox('Choose Jathi', ["Tisra", "Chaturasra", "Khanda", "Misra", "Sankeerna"])
        jathi_map = {"Tisra": 3, "Chaturasra": 4, "Khanda": 5, "Misra": 7, "Sankeerna": 9}
        k = jathi_map[jathi]
        aks = [3*k+2, 2*k+2, k+2, k+3, k+4, 2*k+4, k]
        cycle = next(aks[i] * 4 for i in range(7) if sheet.cell(row=i+6, column=10).value == tala)
    
    if common == "Other":
        tala = st.text_input('Enter name of the Talam')
        cycle = st.number_input('Enter number of Aksharams in one avarthanam of the Talam:', min_value=1, step=1)
    
    eduppu = st.number_input('Enter the Eduppu (Aksharams after Samam where the Corvai is intended to end):', min_value=0, step=1)
    choice = st.selectbox("Place from where the Corvai will START", ["Samam", "Eduppu"]) if lang == "Tamil" else st.selectbox("Place from where the Mugimpu will START", ["Samam", "Jaaga"]) 
    
# Main layout
st.title("MUGIMPU CALCULATOR" if lang == "Telugu" else "CORVAI CALCULATOR")
st.header("INTRODUCTION")
st.write("""Corvai (or Mugimpu in Telugu) is a repeated swara pattern marking the end of a section in a Carnatic concert. This calculator determines the required number of aksharams.""")

st.write(f"Talam is {'Jathi ' + tala if jathi else tala}")
st.write(f"Number of Aksharams in the cycle = {cycle}")
st.write(f"Eduppu aksharams = {eduppu}")
st.write(f"You have chosen to start from {choice} and end at {'Jaaga' if lang == 'Telugu' else 'Eduppu'}")

# Calculations
if choice in ["Eduppu", "Jaaga"]:
    eduppu = 0

a = next((i+1 for i in range(3) if (cycle * (i+1) + eduppu) % 3 == 0), 0)

if a == 0:
    st.error("Not possible. Please check your inputs.")
else:
    n1 = np.array([a + 3 * i for i in range(10)])
    Corvai = (cycle * n1 + eduppu) / 3
    st.write("Possible aksharam values:")
    for i in range(10):
        cols = st.columns(2)
        cols[0].metric('Aksharams per round', value=int(Corvai[i]))
        cols[1].metric('Total Avarthanams', value=int(n1[i]))
