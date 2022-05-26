import numpy as np
import streamlit as st
from openpyxl import load_workbook

workbook = load_workbook(filename="Corvai.xlsx")
sheet=workbook.active

st.set_page_config(layout="wide", page_title="Swara Pattern Calculator")

jathi=0
k=0


with st.sidebar:
    lang=st.selectbox('Select Language of Terminology',["Telugu","Tamil"])
    common=st.selectbox('Choose the Talam',["Adi","Rupakam","Adi Chowkam","Misra Chapu","Khanda Chapu","Adi (Tisra Gathi)","Sapta Talas","Other"])
    for i in range(6):
        if common==sheet.cell(row=i+5,column=1).value:
            cycle=sheet.cell(row=i+5,column=2).value 
            tala=common

    if common=="Sapta Talas":
        tala=st.selectbox('Choose the Talam',["Dhruva","Matya","Rupaka","Jhampa","Triputa","Ata","Eka"])
        jathi=st.selectbox('Choose Jathi',["Tisra","Chaturasra","Khanda","Misra","Sankeerna"])
        if jathi=="Tisra":
            sheet["K2"]=3
            k=3
        elif jathi=="Chaturasra":
            sheet["K2"]=4
            k=4
        elif jathi=="Khanda":
            sheet["K2"]=5
            k=5
        elif jathi=="Misra":
            sheet["K2"]=7
            k=7
        elif jathi=="Sankeerna":
            sheet["K2"]=9
            k=9

        aks=[3*k+2,2*k+2,k+2,k+3,k+4,2*k+4,k]


        for i in range(7):
            if tala==sheet.cell(row=i+6,column=10).value:
                cycle=aks[i]*4


    if common=="Other":
        tala=st.text_input('Enter name of the Talam')
        cycle=st.number_input('Enter number of Aksharams in one avarthanam of the Talam : ',min_value=0,step=1)

    if lang=="Telugu":
        eduppu=st.number_input('Enter the Jaaga (Aksharams after Samam where the Mugimpu is intended to end) :',min_value=0,step=1)
    else:
        eduppu=st.number_input('Enter the Eduppu (Aksharams after Samam where the Corvai is intended to end) :',min_value=0,step=1)

    if lang=="Telugu":
        choice=st.selectbox("Place from where the Mugimpu will START",["Samam","Jaaga"])
    else:
        choice=st.selectbox("Place from where the Corvai will START",["Samam","Eduppu"])


if lang=="Telugu":
    st.title("MUGIMPU CALCULATOR")
    st.header("INTRODUCTION")
    st.write(r"""Mugimpu is one of the most important elements in a Carnatic Vocal concert. Mugimpu refers to a specific swara pattern which is sung *thrice*, and indicates the end of manodharmam for the
particular krithi sung. Mugimpus are generally sung for sub-main and/or main pieces in a concert. Mugimpu is divided into 2 parts, the first part is a swara pattern which is sung thrice in succession, 
followed by the second part which is a combination of aksharams (3,4,5,6,7,8,9, or even 10,11) and gaps. This website calculates the **number of aksharams** a mugimpu needs to be made in, depending on the inputs given by the user """)

else:
    st.title("CORVAI CALCULATOR")
    st.header("INTRODUCTION")
    st.write(r"""Corvai is one of the most important elements in a Carnatic Vocal concert. Corvai refers to a specific swara pattern which is sung *thrice*, and indicates the end of manodharmam for the
particular krithi sung. Corvais are generally sung for sub-main and/or main pieces in a concert. Corvai is divided into 2 parts, the first part is a swara pattern which is sung thrice in succession, 
followed by the second part which is a combination of aksharams (3,4,5,6,7,8,9, or even 10,11) and gaps. This website calculates the **number of aksharams** a corvai needs to be made in, depending on the inputs given by the user """)


if jathi:
    st.write('Talam is ', jathi, "Jathi", tala)
else:
    st.write('Talam is', tala)


st.write("Number of Aksharams in the cycle =",cycle)
if lang=="Telugu":
    st.write("Jaaga aksharams =",eduppu)
else:
    st.write("Eduppu aksharams =",eduppu)

if lang=="Telugu":
    st.write("You have chosen the Mugimpu to start from the", choice ," and end at the Jaaga")
else:
    st.write("You have chosen the Corvai to start from the", choice ," and end at the Eduppu")  

if choice=="Eduppu" or choice=="Jaaga":
    eduppu=0

n=[1,2,3]
a=0
for i in range(3):
    chk=cycle*(i+1)+eduppu
    if chk%3==0:
        a=i+1       
if a==0:
    st.write('Not possible')
    st.write('Please Check.')
    
if a!=0:
    n1=np.zeros(10)
    for i in range(10):
        n1[i]=a+3*i

    Corvai=(cycle*n1+eduppu)/3
    

    if lang=="Telugu":
        st.write("Mugimpu can be made for the following aksharams (Note: This is an arithematic progression, which is an infinite series. Only the first 10 terms have been shown) ")

    else:
        st.write("Corvai can be made for the following aksharams (Note: This is an arithematic progression, which is an infinite series. Only the first 10 terms have been shown) ")

    
    if lang=="Telugu":
        for i in range(10):
            cols=st.columns(2)
            cols[0].metric('Aksharams for one round of Mugimpu',value=np.around(Corvai[i],decimals=0))
            cols[1].metric('Total Avarthanams',value=np.around(n1[i],decimals=0))

    else:
        for i in range(10):
            cols=st.columns(2)
            cols[0].metric('Aksharams for one round of Corvai',value=np.around(Corvai[i],decimals=0))
            cols[1].metric('Total Avarthanams',value=np.around(n1[i],decimals=0))

st.caption("Created by : Sonti Siddharth")

# if jathi=="Tisra":
#             sheet.cell(row=2,column=11)=3
#         elif jathi=="Chaturasra":
#             sheet.cell(row=2,column=11)=4
#         elif jathi=="Khanda":
#             sheet.cell(row=2,column=11)=3
